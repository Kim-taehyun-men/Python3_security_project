import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook

url = 'https://news.v.daum.net/v/20211129144552297'
headers = {
        'User-Agent': 'Mozilla/5.0',
        'Content-Type': 'text/html; charset=utf-8'
        }
response = requests.get(url, headers=headers)
results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)
results = list(set(results))
#set 함수를 이용해 중복을 제거 후 다시 리스트로 저
print(results)

try:
    #email.xlsx 파일을 열고 데이터정보만 가져온다
    wb = load_workbook(r"email.xlsx", data_only=True)
    #활성화하여 sheet 객체로 가져옴
    sheet  = wb.active
except:
    
    #없으면 열고 데이터 삽입
    wb = Workbook()
    sheet = wb.active

    #정보를 입력
for result in results:
    sheet.append([result])

wb.save(r"email.xlsx")
